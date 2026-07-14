#!/usr/bin/env python3
"""
ULTA 대시보드 자동 업데이트 파이프라인
=======================================
매주 ULTA Supplier Portal에서 받는 두 종류의 원본 리포트를 읽어
celimax_ulta 대시보드가 쓰는 data.json 형식으로 재계산합니다.

입력 파일 (파일명 규칙 그대로 유지해야 함):
  - Store-Sales_1002000155_ABSORBLAB-YYYY-MM-DD.xlsx   (매장별 매출)
  - Sales_Inv_Perf__1002000155_ABSORBLAB-YYYY-MM-DD.xlsx (SKU별 매출/재고/OOS)

사용법:
  python3 ulta_update_pipeline.py <원본파일들이_있는_폴더> <기존_data.json_경로> <출력_data.json_경로>

동작 방식:
  - 폴더 안의 모든 Store-Sales_*.xlsx / Sales_Inv_Perf__*.xlsx 파일을 전부 다시 읽어서
    weekly / storeRanking / skuSales / skuOos / financials / qtdYtd 섹션을 처음부터 다시 계산합니다.
    (주차별 원본을 매번 통째로 재계산하므로 파일이 늘어나도 항상 정확합니다)
  - target / marketing / kpiMonthly 등 ULTA 리포트와 무관한 섹션은 기존 data.json 값을 그대로 보존합니다.
  - weekly 항목 중 source가 "manual_table"인 주차(과거에 수기로 입력한 주차)는 덮어쓰지 않고 그대로 둡니다.
"""
import sys
import os
import glob
import json
import datetime
import openpyxl

# ---------- 공통 유틸 ----------

def parse_file_date(filename):
    """파일명 끝의 YYYY-MM-DD 를 파싱. 이 날짜에서 하루를 빼면 실제 주간 종료일(weekEndDate)이 됨."""
    base = os.path.basename(filename)
    base = base.rsplit('.', 1)[0]
    date_str = base[-10:]
    file_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    week_end_date = file_date - datetime.timedelta(days=1)
    return week_end_date.isoformat()


def find_header_row(rows, marker, col=0):
    for i, r in enumerate(rows):
        if r and len(r) > col and r[col] == marker:
            return i
    return None


def num(v):
    """엑셀 셀 값이 숫자면 그대로, 빈 문자열/None/텍스트면 None으로 정규화."""
    if isinstance(v, (int, float)):
        return v
    return None


# ---------- Store-Sales 파싱 ----------

def parse_store_sales(path):
    week_end = parse_file_date(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['StoreSalesReport']
    rows = list(ws.iter_rows(values_only=True))

    fiscal_year, fiscal_week_num = None, None
    for r in rows[:6]:
        if r and r[0] and 'Fiscal Week' in str(r[0]):
            txt = str(r[0]).split(':')[-1].strip()
            fiscal_year, fiscal_week_num = txt[:4], txt[4:]
            break

    hdr_idx = find_header_row(rows, 'Store Number')
    header = rows[hdr_idx]
    data_rows = rows[hdr_idx + 1:]

    # 여러 주가 옆으로 나열된 롤링 리포트일 수 있음 -> 마지막 4개 컬럼 블록(최신 주)만 사용
    # 블록: [Units, Sales, UnitsChg%, SalesChg%] 반복
    n_cols = len(header)
    block_starts = list(range(2, n_cols, 4))
    last_block = block_starts[-1]
    units_col, sales_col = last_block, last_block + 1

    total_row = next(r for r in data_rows if r[1] and str(r[1]).startswith('Total:'))
    com_row = next((r for r in data_rows if r[0] == '0902'), None)  # ULTA.COM = 온라인(COM) 채널
    store_rows = [r for r in data_rows if r[0] not in (None, '') and not str(r[1]).startswith('Total:')]

    total_units = num(total_row[units_col]) or 0
    total_sales = round(num(total_row[sales_col]) or 0, 2)
    com_units = (num(com_row[units_col]) or 0) if com_row else 0
    com_sales = round((num(com_row[sales_col]) or 0) if com_row else 0, 2)
    bm_units = total_units - com_units
    bm_sales = round(total_sales - com_sales, 2)
    store_count = len(store_rows) - (1 if com_row else 0)

    top_stores = []
    for r in store_rows:
        if r[0] == '0902':
            continue  # ULTA.COM은 매장 랭킹에서 제외 (오프라인 매장 랭킹이므로)
        sales = num(r[sales_col]) or 0
        units = num(r[units_col]) or 0
        if sales or units:
            top_stores.append({
                "storeNum": r[0], "storeName": r[1],
                "sales": round(sales, 2), "units": units,
            })

    return {
        "weekEndDate": week_end,
        "fiscalYear": fiscal_year,
        "fiscalWeekNum": fiscal_week_num,
        "totalSales": total_sales,
        "totalUnits": total_units,
        "comSales": com_sales,
        "comUnits": com_units,
        "bmSales": bm_sales,
        "bmUnits": bm_units,
        "storeCount": store_count,
        "topStores": top_stores,
    }


# ---------- Sales_Inv_Perf 파싱 (Last Closed Week 시트: SKU별 매출/재고/OOS) ----------

COL_DESC, COL_COST, COL_PRICE = 2, 8, 9
COL_UNITS, COL_SALES = 10, 13
COL_EOHDOLLARS, COL_WOS = 37, 41
COL_STORECOUNT, COL_STOREOUTS, COL_OOSPCT = 107, 108, 109


def parse_sales_inv_perf_week(path):
    week_end = parse_file_date(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Last Closed Week']
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = find_header_row(rows, 'UPC')
    overall = rows[hdr_idx + 1]
    sku_rows = rows[hdr_idx + 2:]

    skus = []
    for r in sku_rows:
        desc = r[COL_DESC]
        if not desc:
            continue
        sales = num(r[COL_SALES])
        if sales is None:
            continue
        oos_raw = num(r[COL_OOSPCT])
        skus.append({
            "desc": str(desc).strip(),
            "cost": num(r[COL_COST]), "price": num(r[COL_PRICE]),
            "units": num(r[COL_UNITS]), "sales": round(sales, 2),
            "eohUsd": num(r[COL_EOHDOLLARS]), "wos": num(r[COL_WOS]),
            "storeCount": num(r[COL_STORECOUNT]), "storeOuts": num(r[COL_STOREOUTS]),
            "oosPct": (oos_raw * 100) if oos_raw is not None else None,
        })

    overall_sales = num(overall[COL_SALES]) or 0
    overall_eoh = num(overall[COL_EOHDOLLARS]) or 0
    overall_wos = num(overall[COL_WOS])
    return {
        "weekEndDate": week_end,
        "totalSales": round(overall_sales, 2),
        "totalUnits": num(overall[COL_UNITS]),
        "invEohUsd": round(overall_eoh, 2),
        "invWos": round(overall_wos, 1) if overall_wos is not None else None,
        "skus": skus,
    }


def parse_qtd_ytd(path):
    """가장 최근 Sales_Inv_Perf 파일의 Quarter to Date / Year to Date 시트에서 누적치를 그대로 가져옴."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet_name, key in [('Quarter to Date', 'qtd'), ('Year to Date', 'ytd')]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        label = None
        for r in rows[:8]:
            for cell in (r or []):
                if isinstance(cell, str) and (cell.startswith('QTD') or cell.startswith('YTD')):
                    label = cell
        overall_row_idx = None
        for i, r in enumerate(rows):
            for j, cell in enumerate(r or []):
                if cell == 'Overall Result':
                    overall_row_idx = i
        hdr_row_idx = overall_row_idx - 1
        header = rows[hdr_row_idx]
        units_col = next(j for j, v in enumerate(header) if v and 'TOTAL' in str(v) and 'Sales TY' in str(v) and 'Units' in str(v))
        sales_col = next(j for j, v in enumerate(header) if v and 'TOTAL' in str(v) and 'Sales TY' in str(v) and '$' in str(v))
        overall = rows[overall_row_idx]
        sales_val = num(overall[sales_col])
        result[key] = {
            "label": label,
            "totalUnits": num(overall[units_col]),
            "totalSalesUsd": round(sales_val, 2) if sales_val is not None else None,
        }
    return result


# ---------- 월간 롤업 계산 ----------

def month_key(date_iso):
    return date_iso[:7]


def build_store_ranking(store_weeks):
    by_month = {}
    for w in store_weeks:
        mk = month_key(w["weekEndDate"])
        by_month.setdefault(mk, {"totalSales": 0.0, "storeCount": 0, "storeAgg": {}, "weeksIncluded": []})
        agg = by_month[mk]
        agg["totalSales"] += w["totalSales"]
        agg["storeCount"] = max(agg["storeCount"], w["storeCount"])
        agg["weeksIncluded"].append(w["weekEndDate"])
        for s in w["topStores"]:
            key = s["storeNum"]
            e = agg["storeAgg"].setdefault(key, {"storeNum": s["storeNum"], "storeName": s["storeName"], "sales": 0.0, "units": 0})
            e["sales"] += s["sales"]
            e["units"] += s["units"]

    out = {}
    for mk, agg in by_month.items():
        stores = sorted(agg["storeAgg"].values(), key=lambda x: -x["sales"])[:20]
        total = round(agg["totalSales"], 2)
        for s in stores:
            s["sales"] = round(s["sales"], 2)
            s["pctOfTotal"] = round(s["sales"] / total * 100, 2) if total else 0
        out[mk] = {
            "totalSales": total,
            "storeCount": agg["storeCount"],
            "topStores": stores,
            "weeksIncluded": sorted(agg["weeksIncluded"]),
        }
    return out


def build_store_weekly(store_weeks):
    """주차별 매장 순위 스냅샷 (전주 대비 순위 변동 비교용). 월간 롤업과 달리 주 단위를 그대로 보존."""
    by_week = {}
    for w in store_weeks:
        sorted_stores = sorted(w["topStores"], key=lambda x: -x["sales"])[:20]
        ranked = []
        for i, s in enumerate(sorted_stores):
            ranked.append({
                "rank": i + 1, "storeNum": s["storeNum"], "storeName": s["storeName"],
                "sales": round(s["sales"], 2), "units": s["units"],
            })
        by_week[w["weekEndDate"]] = {
            "totalSales": w["totalSales"],
            "storeCount": w["storeCount"],
            "topStores": ranked,
        }
    return by_week


def build_sku_sales(inv_weeks):
    by_month = {}
    for w in inv_weeks:
        mk = month_key(w["weekEndDate"])
        by_month.setdefault(mk, {})
        agg = by_month[mk]
        for sku in w["skus"]:
            e = agg.setdefault(sku["desc"], 0.0)
            agg[sku["desc"]] = e + sku["sales"]

    out = {}
    for mk, agg in by_month.items():
        items = sorted(agg.items(), key=lambda kv: -kv[1])
        total = round(sum(v for _, v in items), 2)
        top10 = items[:10]
        top3_sales = round(sum(v for _, v in items[:3]), 2)
        out[mk] = {
            "totalSales": total,
            "skuCount": len(items),
            "top3Sales": top3_sales,
            "top3Pct": round(top3_sales / total * 100, 1) if total else 0,
            "topSkus": [{"desc": d, "sales": round(v, 2), "pct": round(v / total * 100, 2) if total else 0} for d, v in top10],
        }
    return out


def build_sku_weekly(inv_weeks):
    """주차별 SKU 매출 스냅샷 (전주 대비 비교용). 월간 롤업과 달리 주 단위를 그대로 보존."""
    by_week = {}
    for w in inv_weeks:
        items = []
        for sku in w["skus"]:
            if sku["sales"] is None:
                continue
            items.append({"desc": sku["desc"], "sales": round(sku["sales"], 2), "units": sku["units"]})
        items.sort(key=lambda x: -x["sales"])
        total = round(sum(i["sales"] for i in items), 2)
        by_week[w["weekEndDate"]] = {
            "totalSales": total,
            "skuCount": len(items),
            "skus": items,
        }
    return by_week


def build_sku_oos(inv_weeks):
    by_month = {}
    for w in inv_weeks:
        mk = month_key(w["weekEndDate"])
        by_month.setdefault(mk, {})
        agg = by_month[mk]
        for sku in w["skus"]:
            if sku["oosPct"] is None or sku["storeCount"] is None:
                continue
            e = agg.setdefault(sku["desc"], {"weighted": 0.0, "weight": 0, "maxPct": 0.0, "weeks": 0})
            e["weighted"] += sku["oosPct"] * sku["storeCount"]
            e["weight"] += sku["storeCount"]
            e["maxPct"] = max(e["maxPct"], sku["oosPct"])
            e["weeks"] += 1

    out = {}
    for mk, agg in by_month.items():
        top = []
        overall_weighted, overall_weight = 0.0, 0
        for desc, e in agg.items():
            avg_pct = e["weighted"] / e["weight"] if e["weight"] else 0
            top.append({"desc": desc, "avgOosPct": round(avg_pct, 2), "maxOosPct": round(e["maxPct"], 2), "weeks": e["weeks"]})
            overall_weighted += e["weighted"]
            overall_weight += e["weight"]
        top.sort(key=lambda x: -x["avgOosPct"])
        out[mk] = {
            "overallAvgOosPct": round(overall_weighted / overall_weight, 2) if overall_weight else 0,
            "trackedSkuCount": len(top),
            "topOosSkus": top,
        }
    return out


def build_financials(inv_weeks):
    by_month = {}
    for w in inv_weeks:
        mk = month_key(w["weekEndDate"])
        agg = by_month.setdefault(mk, {"marginNum": 0.0, "marginDen": 0.0, "lostSales": 0.0, "lastWeek": None})
        for sku in w["skus"]:
            price, cost, sales = sku["price"], sku["cost"], sku["sales"]
            if price and cost is not None and sales:
                margin_pct = (price - cost) / price
                agg["marginNum"] += margin_pct * sales
                agg["marginDen"] += sales
            if sku["oosPct"] is not None and sales:
                frac = sku["oosPct"] / 100.0
                if frac < 1:
                    agg["lostSales"] += sales * frac / (1 - frac)
        # 재고(EOH$/WOS)는 시점 스냅샷 -> 그 달의 "마지막 업로드 주차" 값을 사용
        if agg["lastWeek"] is None or w["weekEndDate"] > agg["lastWeek"]["weekEndDate"]:
            agg["lastWeek"] = w

    out = {}
    for mk, agg in by_month.items():
        lw = agg["lastWeek"]
        out[mk] = {
            "marginPct": round(agg["marginNum"] / agg["marginDen"] * 100, 2) if agg["marginDen"] else None,
            "invEohUsd": lw["invEohUsd"],
            "invWos": lw["invWos"],
            "invAsOfWeek": lw["weekEndDate"],
            "lostSalesUsd": round(agg["lostSales"], 2),
        }
    return out


# ---------- 메인 ----------

def run(input_dir, existing_json_path, output_json_path):
    with open(existing_json_path, 'r', encoding='utf-8') as f:
        snapshot = json.load(f)

    store_files = sorted(glob.glob(os.path.join(input_dir, 'Store-Sales_*.xlsx')))
    inv_files = sorted(glob.glob(os.path.join(input_dir, 'Sales_Inv_Perf*.xlsx')))

    print(f"Store-Sales 파일 {len(store_files)}개, Sales_Inv_Perf 파일 {len(inv_files)}개 발견")

    store_weeks = [parse_store_sales(p) for p in store_files]
    inv_weeks = [parse_sales_inv_perf_week(p) for p in inv_files]

    # weekly 섹션: manual_table 주차는 보존, 나머지는 파일 기준으로 갱신
    weekly = snapshot.get("weekly", {})
    for w in store_weeks:
        existing = weekly.get(w["weekEndDate"])
        if existing and existing.get("source") == "manual_table":
            continue
        weekly[w["weekEndDate"]] = {
            "weekEndDate": w["weekEndDate"], "fiscalYear": w["fiscalYear"], "fiscalWeekNum": w["fiscalWeekNum"],
            "totalSales": w["totalSales"], "totalUnits": w["totalUnits"],
            "comSales": w["comSales"], "comUnits": w["comUnits"],
            "bmSales": w["bmSales"], "bmUnits": w["bmUnits"],
            "storeCount": w["storeCount"], "source": "file",
            "uploadedAt": datetime.datetime.utcnow().isoformat() + "Z",
        }
    snapshot["weekly"] = weekly

    if store_weeks:
        snapshot["storeRanking"] = {
            "byMonth": build_store_ranking(store_weeks),
            "uploadedAt": datetime.datetime.utcnow().isoformat() + "Z",
            "note": "Store-Sales 원본 xlsx 파일이 업로드된 주차만 집계됨. 일부 월은 부분 데이터일 수 있음.",
        }
        snapshot["storeWeekly"] = {
            "byWeek": build_store_weekly(store_weeks),
            "note": "Store-Sales 원본 리포트 기준 주간 매장별 매출 순위 (전주 대비 순위 변동 비교용)",
            "uploadedAt": datetime.datetime.utcnow().isoformat() + "Z",
        }

    if inv_weeks:
        snapshot["skuSales"] = {
            "byMonth": build_sku_sales(inv_weeks),
            "note": "Sales_Inv_Perf 원본 리포트의 Last Closed Week 시트 기준 SKU별 매출 합산",
            "uploadedAt": datetime.datetime.utcnow().isoformat() + "Z",
        }
        snapshot["skuWeekly"] = {
            "byWeek": build_sku_weekly(inv_weeks),
            "note": "Sales_Inv_Perf 원본 리포트의 Last Closed Week 시트 기준 주간 SKU별 매출 (전주 대비 비교용)",
            "uploadedAt": datetime.datetime.utcnow().isoformat() + "Z",
        }
        snapshot["skuOos"] = {
            "byMonth": build_sku_oos(inv_weeks),
            "note": "Sales_Inv_Perf 원본 리포트의 Store Out of Stock % 컬럼 기준 (판매 상위 8~10개 SKU만 추적됨, storeCount 가중평균)",
            "uploadedAt": datetime.datetime.utcnow().isoformat() + "Z",
        }
        snapshot["financials"] = {
            "byMonth": build_financials(inv_weeks),
            "note": "ULTA 소매 마진은 (Retail Price - Purch Cost)/Retail Price 기준으로, ULTA 채널 마진 근사치입니다 (Celimax 자체 원가/마진과는 다름). 재고 자산가치/WOS는 해당 월 마지막 업로드 주차 기준 시점 스냅샷입니다. 품절 손실 매출은 OOS%를 이용한 근사 추정치입니다.",
            "uploadedAt": datetime.datetime.utcnow().isoformat() + "Z",
        }

        latest_inv_file = inv_files[-1]
        qtd_ytd = parse_qtd_ytd(latest_inv_file)
        latest_week_end = parse_file_date(latest_inv_file)
        snapshot["qtdYtd"] = {
            "qtd": qtd_ytd["qtd"], "ytd": qtd_ytd["ytd"],
            "asOfDate": latest_week_end,
            "note": "ULTA Sales_Inv_Perf 리포트의 Quarter to Date / Year to Date 시트에서 그대로 가져온 누적 실적입니다.",
        }

    snapshot["snapshotUpdatedAt"] = datetime.datetime.utcnow().isoformat() + "Z"

    with open(output_json_path, 'w', encoding='utf-8') as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=None)

    print(f"완료 -> {output_json_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("사용법: python3 ulta_update_pipeline.py <원본파일_폴더> <기존_data.json> <출력_data.json>")
        sys.exit(1)
    run(sys.argv[1], sys.argv[2], sys.argv[3])
